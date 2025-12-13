"""
Excel Export Service for Extraction Results
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Service for exporting extraction results to Excel format"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def export_extraction_results(
        self,
        documents: List[Dict[str, Any]],
        template_name: str,
        field_names: List[str],
        include_metadata: bool = True
    ) -> BytesIO:
        """
        Export extraction results to Excel file
        
        Args:
            documents: List of document extraction results
            template_name: Name of the template
            field_names: List of field names to export
            include_metadata: Whether to include metadata sheet
            
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        
        # Sheet 1: Extraction Results
        ws_results = wb.active
        ws_results.title = "Extraction Results"
        self._create_results_sheet(ws_results, documents, field_names)
        
        # Sheet 2: Summary (if metadata enabled)
        if include_metadata:
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, documents, template_name)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_results_sheet(
        self,
        ws,
        documents: List[Dict[str, Any]],
        field_names: List[str]
    ):
        """Create the main extraction results sheet"""
        
        # Define headers
        headers = [
            "Document ID",
            "Filename",
            "Status",
            "Upload Date",
            "Validated Date",
        ] + field_names
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Write data
        for row_idx, doc in enumerate(documents, start=2):
            extraction_result = doc.get('extraction_result', {})
            
            # Handle different possible structures of extraction_result
            # Structure 1: {"extracted_data": {...}, "metadata": {...}}
            # Structure 2: Direct field data at root level
            if 'extracted_data' in extraction_result:
                extracted_data = extraction_result.get('extracted_data', {})
                metadata = extraction_result.get('metadata', {})
            else:
                # If no 'extracted_data' key, assume fields are at root level
                extracted_data = extraction_result
                metadata = extraction_result.get('metadata', {})
            
            # Basic info
            ws.cell(row=row_idx, column=1, value=doc.get('id'))
            ws.cell(row=row_idx, column=2, value=doc.get('filename'))
            ws.cell(row=row_idx, column=3, value=doc.get('status', '').upper())
            
            # Dates
            upload_date = doc.get('created_at', '')
            if upload_date:
                ws.cell(row=row_idx, column=4, value=upload_date)
            
            # Validated date is in documents table, not in metadata
            validated_at = doc.get('validated_at', '')
            if validated_at:
                ws.cell(row=row_idx, column=5, value=validated_at)
            
            # Extracted fields - start from column 6 (after basic info columns)
            for idx, field_name in enumerate(field_names):
                col_idx = 6 + idx  # Column 6 onwards for extracted fields
                value = extracted_data.get(field_name, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply borders
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = self.border
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            
            # Check data length
            for row_idx in range(2, len(documents) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width (with some padding)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions
    
    def _create_summary_sheet(
        self,
        ws,
        documents: List[Dict[str, Any]],
        template_name: str
    ):
        """Create summary statistics sheet"""
        
        # Title
        ws.cell(row=1, column=1, value="Export Summary").font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Template info
        ws.cell(row=3, column=1, value="Template:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=template_name)
        
        ws.cell(row=4, column=1, value="Export Date:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Statistics
        ws.cell(row=6, column=1, value="Statistics").font = Font(bold=True, size=12)
        
        total_docs = len(documents)
        validated_docs = sum(1 for doc in documents if doc.get('status') == 'validated')
        pending_docs = sum(1 for doc in documents if doc.get('status') == 'pending')
        
        # Calculate average accuracy
        accuracies = []
        for doc in documents:
            metadata = doc.get('extraction_result', {}).get('metadata', {})
            accuracy = metadata.get('accuracy')
            if accuracy is not None:
                accuracies.append(accuracy)
        
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        
        stats = [
            ("Total Documents", total_docs),
            ("Validated", validated_docs),
            ("Pending", pending_docs),
            ("Average Accuracy", f"{avg_accuracy * 100:.2f}%"),
        ]
        
        for idx, (label, value) in enumerate(stats, start=7):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
